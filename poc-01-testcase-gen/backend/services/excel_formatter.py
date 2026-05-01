from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Column definitions (key, header label) ────────────────────────────────────

_COLUMNS: list[tuple[str, str]] = [
    ("test_id",         "Test ID"),
    ("scenario",        "Scenario"),
    ("preconditions",   "Preconditions"),
    ("steps",           "Steps"),
    ("expected_result", "Expected Result"),
    ("risk_level",      "Risk Level"),
]

# ── Styles ────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_ROW_FILLS: tuple[PatternFill, PatternFill] = (
    PatternFill("solid", fgColor="FFFFFF"),   # even data rows (0-indexed)
    PatternFill("solid", fgColor="DCE6F1"),   # odd data rows
)

_RISK_STYLES: dict[str, tuple[PatternFill, Font]] = {
    "High":   (PatternFill("solid", fgColor="FF0000"), Font(color="FFFFFF")),
    "Medium": (PatternFill("solid", fgColor="FFC000"), Font(color="000000")),
    "Low":    (PatternFill("solid", fgColor="92D050"),  Font(color="000000")),
}

_WRAP = Alignment(wrap_text=True, vertical="top")
_MAX_COL_WIDTH = 60

# Resolved once at module load — avoids magic number in the data loop.
_RISK_COL: int = next(
    i for i, (key, _) in enumerate(_COLUMNS, start=1) if key == "risk_level"
)


# ── Internal helpers ──────────────────────────────────────────────────────────

def _cell_text(value: Any) -> str:
    """Coerce *value* to a cell string; join lists with newlines."""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    return str(value) if value is not None else ""


def _fit_width(ws: Any, col_idx: int) -> float:
    """Return the ideal display width for column *col_idx*, capped at _MAX_COL_WIDTH."""
    max_len = 0
    for cell in ws[get_column_letter(col_idx)]:
        if cell.value:
            text = str(cell.value)
            # Measure the longest line for wrapped multi-line cells.
            line_max = max(len(ln) for ln in text.splitlines()) if "\n" in text else len(text)
            max_len = max(max_len, line_max)
    return min(max_len + 2, _MAX_COL_WIDTH)  # +2 for cell padding


# ── Public API ────────────────────────────────────────────────────────────────

def format(test_cases: list[dict[str, Any]], output_path: str) -> str:  # noqa: A001
    """Write *test_cases* to an Excel workbook at *output_path*.

    Columns in order: Test ID | Scenario | Preconditions | Steps |
    Expected Result | Risk Level.

    Styling:
        - Header row: bold white text on dark blue (#1F3864).
        - Data rows: alternating white (#FFFFFF) and light blue (#DCE6F1).
        - Risk Level cell overrides the row fill: High = red (#FF0000, white
          text), Medium = orange (#FFC000), Low = green (#92D050).
        - Column widths are auto-fitted from content, capped at 60 characters.

    Args:
        test_cases: Non-empty list of dicts. Each dict must contain the six
                    keys produced by ``response_parser.parse()``. Missing keys
                    raise ``KeyError`` and propagate to the caller.
        output_path: Full filesystem path for the ``.xlsx`` file to create.
                     The parent directory must already exist.

    Returns:
        *output_path* unchanged, so callers can chain or log it.

    Raises:
        ValueError: If *test_cases* is empty.
        IOError: If the parent directory of *output_path* does not exist.
        OSError: If openpyxl cannot write the file (permissions, disk full, etc.).
    """
    if not test_cases:
        raise ValueError("test_cases must not be empty")

    output_dir = Path(output_path).parent
    if not output_dir.exists():
        raise IOError(f"Output directory does not exist: {output_dir}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (_, label) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    for data_idx, tc in enumerate(test_cases):
        row = data_idx + 2                    # row 1 is the header
        row_fill = _ROW_FILLS[data_idx % 2]  # 0 → white, 1 → light blue, repeating

        for col_idx, (key, _) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=row, column=col_idx, value=_cell_text(tc[key]))
            cell.fill = row_fill
            cell.alignment = _WRAP

        # Risk Level colour overrides the alternating row fill.
        risk = str(tc.get("risk_level", "Medium")).strip().title()
        if risk in _RISK_STYLES:
            risk_fill, risk_font = _RISK_STYLES[risk]
            risk_cell = ws.cell(row=row, column=_RISK_COL)
            risk_cell.fill = risk_fill
            risk_cell.font = risk_font

    # ── Column widths + freeze ────────────────────────────────────────────────
    for col_idx in range(1, len(_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _fit_width(ws, col_idx)

    ws.freeze_panes = "A2"
    wb.save(output_path)

    logger.info("Wrote %d test case(s) to %s", len(test_cases), Path(output_path).name)
    return output_path
